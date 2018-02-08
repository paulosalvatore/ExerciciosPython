"""
Exercício 28

Nome: Planilhas de Políticos
Objetivo: Desenvolver uma aplicação que busque aleatoriamente alguns políticos e salve determinadas informações em uma planilha de Excel.
Dificuldade: Avançado

1 - Pegar um político na API http://politicos.olhoneles.org/api/v0/politicians/1/
2 - Armazenar todas as informações desse político em uma planilha de Excel seguindo esse modelo: http://gg.gg/excel_politicos
3 - Depois de pronto, faça o código pegar 10 números de ID aleatoriamente e gravar uma worksheet para cada político.
4 - O nome de cada página deve ser o nome do político e as páginas deverão estar ordenadas por ordem alfabética.
"""

# Resolução

from openpyxl import Workbook
import requests
import json
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
import io
import urllib3
from random import randint
import time


tempo_inicio = time.clock()


def pegar_valor_dicionario(caminho, base):
	try:
		for item in caminho:
			base = base[item]
	except:
		base = "Valor não informado"
	return base


def colocar_valores(linha_inicial, coluna_inicial, informacoes):
	linha_atual = linha_inicial

	coluna_titulo = coluna_inicial
	coluna_valor = coluna_titulo + 1

	for informacao in informacoes:
		if informacao.get("centralizar"):
			ws.cell(row=linha_atual, column=coluna_titulo).alignment = centralizado
			ws.cell(row=linha_atual, column=coluna_valor).alignment = centralizado

		ws.cell(
			row=linha_atual,
			column=coluna_titulo,
			value=informacao["titulo"] + ":"
		)

		ws.cell(row=linha_atual, column=coluna_titulo).font = fonte_negrito

		valor = informacao.get("valor")

		if valor == None:
			ws.merge_cells(
				start_row=linha_atual,
				end_row=linha_atual,
				start_column=coluna_titulo,
				end_column=coluna_valor
			)

			linha_atual += 1
		else:
			if type(valor) == list:
				for item in valor:
					ws.cell(
						row=linha_atual,
						column=coluna_valor,
						value=item
					)

					linha_atual += 1
			else:
				ws.cell(
					row=linha_atual,
					column=coluna_valor,
					value=valor
				)

				linha_atual += 1


maximo_politicos = 1438992

lista_politicos = []

while len(lista_politicos) < 10:
	politico_aleatorio = randint(1, maximo_politicos)
	if lista_politicos.count(politico_aleatorio) == 0:
		lista_politicos.append(politico_aleatorio)

wb = Workbook()

for indice_politico, politico in enumerate(lista_politicos):
	url = "http://politicos.olhoneles.org/api/v0/politicians/{}".format(politico)
	conteudo_url = requests.get(url)
	conteudo_json = json.loads(conteudo_url.text)

	if indice_politico == 0:
		ws = wb.active
	else:
		ws = wb.create_sheet()

	ws.title = conteudo_json["alternative_names"][0]["name"]

	nome = pegar_valor_dicionario(["name"], conteudo_json)
	cpf = pegar_valor_dicionario(["cpf"], conteudo_json)
	data_nascimento = pegar_valor_dicionario(["date_of_birth"], conteudo_json)
	genero = pegar_valor_dicionario(["gender"], conteudo_json)
	estado_civil = pegar_valor_dicionario(["marital_status", "name"], conteudo_json)
	educacao = pegar_valor_dicionario(["education", "name"], conteudo_json)
	etnia = pegar_valor_dicionario(["ethnicity", "name"], conteudo_json)
	nacionalidade = pegar_valor_dicionario(["nationality", "name"], conteudo_json)
	estado = pegar_valor_dicionario(["state", "name"], conteudo_json)
	ocupacao = pegar_valor_dicionario(["occupation", "name"], conteudo_json)
	local_nascimento = pegar_valor_dicionario(["place_of_birth"], conteudo_json)
	candidaturas = pegar_valor_dicionario(["candidacies"], conteudo_json)
	partidos = pegar_valor_dicionario(["political_parties"], conteudo_json)

	for indice, partido in enumerate(partidos):
		partido = partido["political_party"]
		exibicao = "{} - {} ({})".format(partido["siglum"], partido["name"], partido["tse_number"])
		partidos[indice] = exibicao

	informacoes_politico = [
		{
			"titulo": "Informações do Político",
			"centralizar": True
		},
		{
			"titulo": "Nome",
			"valor": nome
		},
		{
			"titulo": "CPF",
			"valor": cpf
		},
		{
			"titulo": "Data de Nascimento",
			"valor": data_nascimento
		},
		{
			"titulo": "Gênero",
			"valor": genero
		},
		{
			"titulo": "Estado Civil",
			"valor": estado_civil
		},
		{
			"titulo": "Educação",
			"valor": educacao
		},
		{
			"titulo": "Etnia",
			"valor": etnia
		},
		{
			"titulo": "Nacionalidade",
			"valor": nacionalidade
		},
		{
			"titulo": "Estado",
			"valor": estado
		},
		{
			"titulo": "Ocupação",
			"valor": ocupacao
		},
		{
			"titulo": "Local Nascimento",
			"valor": local_nascimento
		},
		{
			"titulo": "Partidos",
			"valor": partidos
		}
	]

	informacoes_candidatura = [
		{
			"titulo": "Candidaturas",
			"centralizar": True
		}
	]

	for indice, candidatura in enumerate(candidaturas):
		lista_candidatura = [
			{
				"titulo": "{}ª Candidatura".format(indice + 1)
			},
			{
				"titulo": "Cargo",
				"valor": pegar_valor_dicionario(["political_parties", "name"], candidatura)
			},
			{
				"titulo": "Status",
				"valor": pegar_valor_dicionario(["candidacy_status", "name"], candidatura)
			},
			{
				"titulo": "Eleito",
				"valor": pegar_valor_dicionario(["elected"], candidatura)
			},
			{
				"titulo": "Cidade",
				"valor": pegar_valor_dicionario(["city", "name"], candidatura)
			},
			{
				"titulo": "Estado",
				"valor": pegar_valor_dicionario(["state", "name"], candidatura)
			}
		]

		for item in lista_candidatura:
			informacoes_candidatura.append(item)

	fonte_negrito = Font(bold=True)
	centralizado = Alignment(horizontal="center")

	colocar_valores(2, 4, informacoes_politico)
	colocar_valores(2, 7, informacoes_candidatura)

	try:
		http = urllib3.PoolManager()
		foto = conteudo_json["picture"]
		request_foto = http.request("GET", foto)
		arquivo_imagem = io.BytesIO(request_foto.data)
		img = Image(arquivo_imagem)
		ws.add_image(img, "B2")
		ws.column_dimensions["B"].width = img.drawing.width / 7
	except:
		pass

	ws.column_dimensions["C"].width = 3
	ws.column_dimensions["F"].width = 3

	largura_minima_coluna = 10

	dims = {}

	for row in ws.rows:
		for cell in row:
			if cell.value:
				dims[cell.column] = max(
					dims.get(cell.column, 0),
					len(str(cell.value)) + 1
				)

	for col, value in dims.items():
		ws.column_dimensions[col].width = max(value, largura_minima_coluna)

	print("Página '{}' gerada.".format(ws.title))

wb._sheets.sort(key=lambda ws: ws.title)

wb.save("planilha.xlsx")

tempo_final = time.clock() - tempo_inicio

print("Tempo demorado: {:.2f} segundos.".format(tempo_final))
